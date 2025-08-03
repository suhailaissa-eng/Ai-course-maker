import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def clean_illegal_chars(text):
    if not isinstance(text, str):
        return text

    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\uFFFE\uFFFF]', '', text)

def structure_block_segments(block_title, segments):
    types = ["Video", "Reading", "Assignment", "Quiz", "Discussion"]
    video_types = ["Talking head", "Light Board", "Screencast", "Lab Interview"]
    result = []
    video_count = 0

    for i, seg in enumerate(segments[:7]):
        t = types[i % len(types)]
        entry = {
            "segment_title": seg["title"],
            "learning_type": t,
            "includes": [seg["title"]]
        }
        if t == "Video":
            entry["video_type"] = video_types[video_count % len(video_types)]
            video_count += 1
        result.append(entry)
    return result

def export_block_excel(module, block, structured, out_dir="uploads/review"):
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{module}_{block}.xlsx".replace(" ", "_"))
    rows = []

    for i, seg in enumerate(structured):
        for j, item in enumerate(seg["includes"]):
            rows.append({
                "Module": module if i == 0 and j == 0 else "",
                "Block": block if i == 0 and j == 0 else "",
                "Learning Segment Title": item,
                "Learning Type": seg["learning_type"],
                "Video Type": seg.get("video_type", "")
            })
    pd.DataFrame(rows).to_excel(path, index=False)
    return path

def validate_edited_excel(path):
    try:
        df = pd.read_excel(path)
        required = ["Module", "Block", "Learning Segment Title", "Learning Type", "Video Type"]
        for col in required:
            if col not in df.columns:
                raise ValueError(f"Missing column: {col}")
        return df
    except Exception as e:
        print(f"Validation error: {e}")
        return None

def export_final_excel(data, path):

    rows = []
    for d in data:
        module, block = d["module"], d["block"]
        for seg in d["structured_segments"]:
            for s in seg["includes"]:
                rows.append({
                    "Module (LOs)": clean_illegal_chars(module),
                    "Blocks (Learning Weeks)": clean_illegal_chars(block),
                    "Learning Segment Title": clean_illegal_chars(s),
                    "Learning Segment Type": clean_illegal_chars(seg["learning_type"]),
                    "Video Type": clean_illegal_chars(seg.get("video_type", "")),
                    "Video Link": ""
                })

    wb = Workbook()
    ws = wb.active
    ws.title = "Course Outline"

    headers = [
        "Module (LOs)",
        "Blocks (Learning Weeks)",
        "Learning Segment Title",
        "Learning Segment Type",
        "Video Type",
        "Video Link"
    ]
    ws.append(headers)

    for row in rows:
        ws.append([clean_illegal_chars(row[h]) for h in headers])

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [20, 25, 35, 25, 15, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    def merge_similar_cells(col_idx):
        start = 2
        prev = ws.cell(row=2, column=col_idx).value
        for row in range(3, ws.max_row + 2):
            curr = ws.cell(row=row, column=col_idx).value
            if curr != prev or curr is None:
                if row - start > 1:
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                    for r in range(start, row):
                        ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="center", horizontal="center")
                start = row
                prev = curr

        if ws.max_row + 1 - start > 0:
            ws.merge_cells(start_row=start, start_column=col_idx, end_row=ws.max_row + 1, end_column=col_idx)
            for r in range(start, ws.max_row + 2):
                ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="center", horizontal="center")

    merge_similar_cells(1)  
    merge_similar_cells(2)  

    
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        fill = fill1 if row % 2 == 0 else fill2
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill

    wb.save(path)
