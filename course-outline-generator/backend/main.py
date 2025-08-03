from fastapi import FastAPI, File, UploadFile, Body, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict
import os, shutil
from sentence_transformers import SentenceTransformer, util
from parsers import extract_modules_blocks_from_docx, parse_pptx, parse_pdf, parse_docx
from utils import structure_block_segments, export_block_excel, validate_edited_excel
import pandas as pd
from docx import Document
import fitz  
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
model = SentenceTransformer("all-MiniLM-L6-v2")
all_block_outputs = []

@app.post("/generate-outline")
async def generate_outline(
    plan: UploadFile = File(...),
    slides: List[UploadFile] = File([]),
    assignments: List[UploadFile] = File([])
):
    plan_path = os.path.join(UPLOAD_DIR, plan.filename)
    with open(plan_path, "wb") as f:
        shutil.copyfileobj(plan.file, f)

    module_data = extract_modules_blocks_from_docx(plan_path)
    all_segments = []

    for slide in slides:
        ext = os.path.splitext(slide.filename)[-1].lower()
        path = os.path.join(UPLOAD_DIR, slide.filename)
        with open(path, "wb") as f:
            shutil.copyfileobj(slide.file, f)

        if ext == ".pptx":
            all_segments.extend(parse_pptx(path))
        elif ext == ".pdf":
            all_segments.extend(parse_pdf(path))
        elif ext == ".docx":
            all_segments.extend(parse_docx(path))

    global all_block_outputs
    all_block_outputs = []

    for entry in module_data:
        module = entry["module"]
        for block in entry["blocks"]:
            block_emb = model.encode(block, convert_to_tensor=True)
            scored = []
            for seg in all_segments:
                combined = f"{seg['title']} {seg['summary']}"
                emb = model.encode(combined, convert_to_tensor=True)
                score = util.cos_sim(block_emb, emb).item()
                scored.append((score, seg))

            top = sorted(scored, key=lambda x: x[0], reverse=True)[:15]
            top_segs = [s for _, s in top]

            structured = structure_block_segments(block, top_segs)
            excel_path = export_block_excel(module, block, structured)
            all_block_outputs.append({
                "module": module,
                "block": block,
                "structured_segments": structured
            })

    formatted_modules = []
    for entry in all_block_outputs:
        mod = next((m for m in formatted_modules if m["module_title"] == entry["module"]), None)
        if not mod:
            mod = {"module_title": entry["module"], "blocks": []}
            formatted_modules.append(mod)
        mod["blocks"].append({
            "block_title": entry["block"],
            "segments": entry["structured_segments"]
        })

    return {"modules": formatted_modules}

@app.post("/finalize-outline")
async def finalize_outline(selected_modules: List[Dict] = Body(...)):
    try:
        output_path = os.path.join(UPLOAD_DIR, "final_course_outline.xlsx")
        
        all_data = []
        for mod in selected_modules:
            module_title = mod["module_title"]
            for block in mod["blocks"]:
                block_title = block["block_title"]
                structured_segments = []
                for seg in block["segments"]:
                    structured_segments.append({
                        "includes": [seg["segment_title"]],
                        "learning_type": seg["learning_type"],
                        "video_type": seg.get("video_type", "")
                    })
                all_data.append({
                    "module": module_title,
                    "block": block_title,
                    "structured_segments": structured_segments
                })

        export_final_excel(all_data, output_path)

        return {
            "message": "Final Excel outline created!",
            "download_url": "/download-final-excel"
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")
    
@app.get("/download-final-excel")
async def download_final_excel():
    path = os.path.join(UPLOAD_DIR, "final_course_outline.xlsx")
    return FileResponse(
        path,
        filename="Final_Course_Outline.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



def clean_illegal_chars(text):
    import re
    if not isinstance(text, str):
        return text
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\uFFFE\uFFFF]', '', text)

def export_final_excel(data, path):
    rows = []
    for d in data:
        module, block = d["module"], d["block"]
        for seg in d["structured_segments"]:
            for s in seg["includes"]:
                rows.append({
                    "Module (LOs)": clean_illegal_chars(module),
                    "Blocks (Learning Weeks)": clean_illegal_chars(block),
                    "Learning Segment Title": clean_illegal_chars(s),
                    "Learning Segment Type": clean_illegal_chars(seg["learning_type"]),
                    "Video Type": clean_illegal_chars(seg.get("video_type", "")),
                    "Video Link": ""
                })

    wb = Workbook()
    ws = wb.active
    ws.title = "Course Outline"

   
    ws["A1"] = "Course Title"
    ws["D1"] = "Launching Date"
    ws["A2"] = "Course SMEs"
    ws["D2"] = "Platform"

    header_fill_pink = PatternFill("solid", fgColor="F4CCCC") 
    header_fill_yellow = PatternFill("solid", fgColor="FFF2CC") 
    header_fill_blue = PatternFill("solid", fgColor="DDEBF7")  
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid") 
    gray1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
    gray2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 

    header_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ["A1", "D1", "A2", "D2"]:
        ws[cell].fill = header_fill_pink
        ws[cell].font = header_font
        ws[cell].alignment = center_align
        ws[cell].border = thin_border

   
    ws.merge_cells("B1:C1")
    ws.merge_cells("B2:C2")
    ws.merge_cells("E1:F1")
    ws.merge_cells("E2:F2")

    for cell in ["B1", "C1", "B2", "C2", "E1", "F1", "E2", "F2"]:
        c = ws[cell]
        c.font = normal_font
        c.alignment = center_align
        c.border = thin_border

   
    headers = [
        "Module (LOs)",
        "Blocks (Learning Weeks)",
        "Learning Segment Title",
        "Learning Segment Type",
        "Video Type",
        "Video Link"
    ]

    ws.append(headers)
    for col_idx, cell_value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        if col_idx == 6:
            cell.fill = header_fill_blue
        else:
            cell.fill = header_fill_yellow

    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border

            
            if col_idx == 1:
                cell.fill = yellow_fill
            else:
                cell.fill = gray1 if row_idx % 2 == 0 else gray2

    col_widths = [20, 25, 35, 25, 15, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    def merge_similar_cells(col_idx):
        start = 4
        prev = ws.cell(row=start, column=col_idx).value
        for row in range(start+1, ws.max_row + 2):
            curr = ws.cell(row=row, column=col_idx).value
            if curr != prev or curr is None:
                if row - start > 1:
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=row-1, end_column=col_idx)
                    for r in range(start, row):
                        c = ws.cell(row=r, column=col_idx)
                        c.alignment = center_align
                        c.border = thin_border
                start = row
                prev = curr
        if ws.max_row + 1 - start > 0:
            ws.merge_cells(start_row=start, start_column=col_idx, end_row=ws.max_row + 1, end_column=col_idx)
            for r in range(start, ws.max_row + 2):
                c = ws.cell(row=r, column=col_idx)
                c.alignment = center_align
                c.border = thin_border

    merge_similar_cells(1)
    merge_similar_cells(2)

    wb.save(path)
